import openpyxl
#
# # 将excle表添加到工作簿
# workbook = openpyxl.load_workbook(filename=r"C:\Projects\Python_01\py_01_day26project\datas\case接口.xlsx")
# # 选择表单
# sh = workbook["login"]
#
# res = list(sh.rows)
# print(res)

# 写入表单的操作
# res = sh.cell(row=15, column=15, value="python")

class HandleExcle:
    def __init__(self, filename, sheetname):
        """

        :param filename: 文件名
        :param sheetname: 表单名
        """
        self.filename = filename
        self.sheetname = sheetname
    def read_data(self):
        # 读取数据
        # 将excle表添加到工作簿r
        workbook = openpyxl.load_workbook(self.filename)
        # 选择表单
        sh = workbook[self.sheetname]
        # 获取数据
        res = list(sh.rows)
        # 获取表头
        title = [i.value for i in res[0]]
        # print(title)
        cases = []
        for item in res[1:]:
            # 获取除表头以外的其他行
            data = [i.value for i in item]
            # 使用zip聚合打包表头和当前行
            dic = dict(zip(title, data))
            cases.append(dic)
        return cases
    def write_data(self, row, column, value):
        # 将excle表添加到工作簿
        workbook = openpyxl.load_workbook(self.filename)
        # 选择表单
        sh = workbook[self.sheetname]
        # 写入表单的操作
        sh.cell(row=row, column=column, value=value)
        workbook.save(self.filename)

if __name__ == '__main__':
    excle = HandleExcle(r"C:\Projects\Python_01\py_01_day26project\datas\case接口.xlsx", "login")
    res = excle.read_data()
print(res)
# 恢复原来的版本

